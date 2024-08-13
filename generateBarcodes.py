import os
import hashlib
from barcode import Code128
from barcode.writer import ImageWriter
import openpyxl
import xlsxwriter

# Create the 'barcodes' directory if it doesn't exist
if not os.path.exists('barcodes'):
    os.makedirs('barcodes')
print("Created 'barcodes' directory.")

# Load the workbook and select the first sheet
print("Loading 'Products.xlsx'...")
old_wb = openpyxl.load_workbook('Products.xlsx')
old_sheet = old_wb.active
print("Loaded 'Products.xlsx'.")

# Create a new workbook and select the first sheet
print("Creating 'Barcodes.xlsx'...")
new_wb = xlsxwriter.Workbook('Barcodes.xlsx')
new_sheet = new_wb.add_worksheet()

# Write the headers
new_sheet.write('A1', 'Product Description')
new_sheet.write('B1', 'Product ID')
new_sheet.write('C1', 'Barcode')
print("Wrote headers to 'Barcodes.xlsx'.")

# Function to generate a 9-digit hash
def generate_hash(input_string):
    hash_object = hashlib.sha256(input_string.encode())
    hex_dig = hash_object.hexdigest()
    return str(int(hex_dig, 16))[:9]

# Iterate over the product descriptions and IDs in the old sheet
for i, row in enumerate(old_sheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True), start=1):
    product_description = row[0]
    product_id = row[1]
    print(f"Processing product ID: {product_id}")

    # Generate the 9-digit hash
    hashed_id = generate_hash(product_id)
    print(f"Generated 9-digit hash: {hashed_id}")

    # Generate the barcode
    barcode = Code128(hashed_id, writer=ImageWriter())

    # Save the barcode to a file with the original product ID as the name
    barcode_filename = f'barcodes/{product_id}.png'
    with open(barcode_filename, 'wb') as f:
        barcode.write(f)
    print(f"Saved barcode to '{barcode_filename}'.")

    # Write the product description, ID, and the hashed ID to the new sheet
    new_sheet.write(i, 0, product_description)
    new_sheet.write(i, 1, product_id)
    new_sheet.write(i, 2, hashed_id)
    print(f"Wrote product description, ID, and hashed ID to 'Barcodes.xlsx'.")

    # Insert the barcode image to the new sheet
    new_sheet.insert_image(i, 3, barcode_filename)
    print(f"Inserted barcode image to 'Barcodes.xlsx'.")

# Close the new workbook
new_wb.close()
print("Closed 'Barcodes.xlsx'.")
